import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from copy import copy


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from copy import copy


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from copy import copy


def RozszerzenieKolumn(path):
    wb = openpyxl.load_workbook(path)
    for ws in wb:
        # Zablokowanie pierwszego wiersza
        ws.freeze_panes = "A2"

        # Formatowanie pierwszego wiersza (nagłówków)
        for cell in ws[1]:
            cell.font = Font(size=12, bold=True)
            cell.fill = PatternFill(fgColor="B8CCE4", fill_type="solid")

        # Gruby obramowanie dla komórek (opcjonalnie, zostawiam, ale nie zmieniam)
        bd = Side(style="thick", color="000000")

        # Automatyczne dostosowanie szerokości kolumn
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Formatowanie kolumn K i N (dynamiczne formatowanie z separatorem tysięcznym)
        for row in range(2, ws.max_row + 1):  # Pomijamy pierwszy wiersz (nagłówki)
            for col in ["K", "N"]:
                cell = ws[f"{col}{row}"]
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    # Sprawdzanie, czy liczba ma wartości po przecinku różne od zera
                    if cell.value == int(cell.value):
                        # Liczba całkowita, brak miejsc po przecinku
                        formatted_value = "{:,.0f}".format(cell.value).replace(",", " ")
                    else:
                        # Liczba z miejscami po przecinku
                        formatted_value = (
                            "{:,.2f}".format(cell.value)
                            .replace(",", " ")
                            .replace(".", ",")
                        )

                    # Aktualizacja wartości komórki jako sformatowanego tekstu
                    cell.value = formatted_value

        # Ukrycie kolumn A, B i E
        for col in ["A", "B", "E"]:
            ws.column_dimensions[col].hidden = True

        # Wyrównanie do środka wszystkich komórek
        for col in ws.columns:
            for cell in col:
                alignment_obj = copy(cell.alignment)
                alignment_obj.horizontal = "center"
                alignment_obj.vertical = "center"
                cell.alignment = alignment_obj

    # Zapisanie zmian
    wb.save(path)
    wb.close()
